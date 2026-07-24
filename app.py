import io
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


st.set_page_config(
    page_title="Validación OT",
    page_icon="📋",
    layout="wide",
)

APP_DIR = Path(__file__).resolve().parent
LOGO_PATH = APP_DIR / "Finning-CAT.png"
OK_COLOR = "#FFC400"
OBS_COLOR = "#EF4444"


# ============================================================
# Estilos visuales
# ============================================================
st.markdown(
    """
    <style>
    .stApp { background: #f6f7f9; }
    .main-title {font-size: 34px; font-weight: 900; color:#111; margin-bottom:0px;}
    .subtitle {font-size: 16px; color:#5b6470; margin-top:0px;}
    section[data-testid="stSidebar"] {background: linear-gradient(180deg,#090909 0%,#1a1a1a 100%);}
    section[data-testid="stSidebar"] * {color: #fff;}
    section[data-testid="stSidebar"] div[data-testid="stImage"] {background:#ffffff; border-radius:6px; padding:4px; border:1px solid #ffffff55;}
    .step-card {border:1px solid #444; border-radius:10px; padding:14px; margin-bottom:14px; background:#141414;}
    .step-number {display:inline-block; background:#ffc400; color:#111 !important; border-radius:50%; width:28px; height:28px; text-align:center; font-weight:900; line-height:28px; margin-right:8px;}
    .kpi-card {background:white; border:1px solid #e1e5ea; border-radius:14px; padding:18px 20px; box-shadow:0 1px 4px rgba(0,0,0,0.04); min-height:115px;}
    .kpi-title {font-size:13px; color:#111; font-weight:800; text-transform:uppercase;}
    .kpi-value {font-size:34px; color:#111; font-weight:900; line-height:1.1;}
    .kpi-note {font-size:13px; color:#657080;}
    .panel {background:white; border:1px solid #e1e5ea; border-radius:14px; padding:18px; box-shadow:0 1px 4px rgba(0,0,0,0.04);}
    .critical-note {background:#fff1f1; border-left:5px solid #dc2626; padding:11px 14px; border-radius:5px; margin:8px 0 15px 0; color:#7f1d1d;}
    div[data-testid="stDownloadButton"] button, div.stButton > button {background:#ffc400; color:#111; border:none; font-weight:800; border-radius:8px;}
    div[data-testid="stFileUploader"] {border:1px dashed #ffc400; border-radius:10px; padding:8px;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# Configuración del formato OT entregado
# ============================================================
BLANK_TOKENS = {
    "",
    "-",
    "--",
    "N/A",
    "NA",
    "NAN",
    "NONE",
    "NULL",
    "S/I",
    "SIN INFORMACION",
    "SIN INFORMACIÓN",
}

# Caracteres invisibles que aparecen en el formulario como relleno de celdas.
INVISIBLE_CHARS = ("\xa0", "\u200b", "\u200c", "\u200d", "\ufeff", "\u3164")

PRIORITY_CRITICAL = "Crítico"
PRIORITY_STANDARD = "Estándar"

CRITICAL_FIELDS = {
    "Código trabajo",
    "Código síntoma",
    "Código causa",
    "Firma jefe turno (nombre + RUT)",
    "Firma técnico responsable (nombre + RUT)",
}

# Orden utilizado en tablas y gráficos de cumplimiento.
VALIDATED_FIELDS = [
    "Horómetro",
    "Motivo detención del equipo",
    "Descripción del síntoma",
    "Código trabajo",
    "Código síntoma",
    "Código causa",
    "Descripción de actividades",
    "Firma jefe turno (nombre + RUT)",
    "Firma técnico responsable (nombre + RUT)",
]

# Códigos de causa clasificados como "Otros" y, por lo tanto, no aceptados.
INVALID_CAUSE_CODES = {6.6, 7.1}

# Filas de captura del formato. Cada registro ocupa un bloque de filas combinadas.
WORK_ROWS = list(range(42, 94, 4))
MECHANIC_ROWS = list(range(145, 164, 3))
SIMS_ROWS = [189, 193, 197]

# Todas las columnas del bloque se usan para detectar qué filas contienen actividad.
WORK_FIELDS: List[Tuple[str, str]] = [
    ("B", "Hora inicio"),
    ("F", "Hora término"),
    ("J", "N° orden servicio"),
    ("Q", "Código componente SMCS"),
    ("T", "Código modificador"),
    ("W", "Código trabajo"),
    ("Z", "Descripción del síntoma"),
    ("AO", "Código síntoma"),
    ("AR", "Descripción de la causa"),
    ("BH", "Código causa"),
    ("BO", "Tipo tarea"),
    ("BV", "Tarea principal"),
]

# Únicos campos exigidos dentro del bloque Información del trabajo.
ESSENTIAL_WORK_FIELDS: List[Tuple[str, str]] = [
    ("W", "Código trabajo"),
    ("Z", "Descripción del síntoma"),
    ("AO", "Código síntoma"),
    ("BH", "Código causa"),
]

MECHANIC_FIELDS: List[Tuple[str, str]] = [
    ("B", "Código mecánico"),
    ("H", "Nombre y apellido trabajador"),
    ("Z", "Orden de servicio mecánico"),
    ("AD", "Horas de trabajo real"),
]

SIMS_FIELDS: List[Tuple[str, str]] = [
    ("B", "N° pieza que falló"),
    ("E", "Descripción de la pieza"),
    ("X", "Cantidad"),
    ("AA", "Código servicio"),
    ("AE", "N° grupo"),
    ("AJ", "Descripción del grupo"),
    ("AR", "Fin de vida útil Sí/No"),
    ("AU", "Comentarios SIMS"),
]

SUMMARY_COLUMNS = [
    "Archivo",
    "Equipo",
    "Orden",
    "Turno",
    "Campos faltantes",
    "Faltantes críticos",
    "Estado",
    "Campos con observación",
]

DETAIL_COLUMNS = [
    "Archivo",
    "Equipo",
    "Orden",
    "Sección",
    "Campo faltante",
    "Prioridad",
    "Fila OT",
    "Celda/Rango",
    "Observación",
]


# ============================================================
# Utilidades de lectura y validación
# ============================================================
def normalize_value(value: Any) -> str:
    """Convierte el valor en texto útil y elimina rellenos invisibles del formulario."""
    if value is None:
        return ""

    text = str(value)
    for char in INVISIBLE_CHARS:
        text = text.replace(char, " ")

    text = " ".join(text.split()).strip()
    return "" if text.upper() in BLANK_TOKENS else text


def is_filled(value: Any) -> bool:
    return normalize_value(value) != ""


def is_zero_numeric(value: Any) -> bool:
    """Retorna True cuando el valor corresponde numéricamente a cero."""
    text = normalize_value(value).replace(",", ".")
    if not text:
        return False

    try:
        return abs(float(text)) < 1e-9
    except (TypeError, ValueError):
        return False


def is_invalid_cause_code(value: Any) -> bool:
    """Retorna True cuando el código causa corresponde a las categorías Otros 6.6 o 7.1."""
    text = normalize_value(value).replace(",", ".")
    if not text:
        return False
    try:
        numeric_value = float(text)
    except (TypeError, ValueError):
        return False
    return any(abs(numeric_value - invalid) < 1e-9 for invalid in INVALID_CAUSE_CODES)


def summarize_observations(missing: Sequence[Dict[str, str]]) -> str:
    """Genera una descripción breve y sin duplicados para el resumen por OT."""
    if not missing:
        return ""

    summaries: List[str] = []
    seen = set()
    for item in missing:
        field = item.get("Campo faltante", "Campo")
        observation = item.get("Observación", "")
        if observation.startswith("Valor inválido"):
            label = f"{field} ({observation})"
        elif observation.startswith("Falta "):
            label = f"{field} ({observation.lower()})"
        else:
            label = field
        if label not in seen:
            summaries.append(label)
            seen.add(label)

    return "; ".join(summaries)


def has_x(value: Any) -> bool:
    return normalize_value(value).upper() == "X"


def get_value(ws, cell: str) -> Any:
    try:
        return ws[cell].value
    except Exception:
        return None


def first_filled(ws, cells: Sequence[str]) -> str:
    for cell in cells:
        value = get_value(ws, cell)
        if is_filled(value):
            return normalize_value(value)
    return ""


def count_x(ws, cells: Sequence[str]) -> int:
    return sum(1 for cell in cells if has_x(get_value(ws, cell)))


def field_priority(field_name: str) -> str:
    return PRIORITY_CRITICAL if field_name in CRITICAL_FIELDS else PRIORITY_STANDARD


def add_missing(
    missing: List[Dict[str, str]],
    section: str,
    field: str,
    cell_range: str,
    observation: str = "Campo requerido sin información",
    row_label: str = "",
    priority: Optional[str] = None,
) -> None:
    missing.append(
        {
            "Sección": section,
            "Campo faltante": field,
            "Prioridad": priority or field_priority(field),
            "Fila OT": row_label,
            "Celda/Rango": cell_range,
            "Observación": observation,
        }
    )


def validate_single_cells(
    ws,
    missing: List[Dict[str, str]],
    section: str,
    fields: Iterable[Tuple[str, str]],
) -> None:
    for field, cell in fields:
        if not is_filled(get_value(ws, cell)):
            add_missing(missing, section, field, cell)


def validate_exclusive_x(
    ws,
    missing: List[Dict[str, str]],
    section: str,
    field: str,
    cells: Sequence[str],
    labels: Sequence[str],
) -> None:
    selected = count_x(ws, cells)
    if selected == 0:
        add_missing(
            missing,
            section,
            field,
            "/".join(cells),
            f"Debe marcar X en una opción: {', '.join(labels)}",
        )
    elif selected > 1:
        add_missing(
            missing,
            section,
            field,
            "/".join(cells),
            f"Existe más de una opción marcada. Debe seleccionar solo una: {', '.join(labels)}",
        )


def active_rows(ws, rows: Sequence[int], columns: Sequence[str]) -> List[int]:
    return [
        row
        for row in rows
        if any(is_filled(get_value(ws, f"{column}{row}")) for column in columns)
    ]


def validate_required_table(
    ws,
    missing: List[Dict[str, str]],
    section: str,
    rows: Sequence[int],
    required_fields: Sequence[Tuple[str, str]],
    trigger_columns: Optional[Sequence[str]] = None,
    require_one_row: bool = True,
    row_name: str = "Registro",
) -> List[int]:
    """Valida solo las filas utilizadas. Si no existe ninguna, valida la primera fila visible."""
    trigger_columns = list(trigger_columns or [column for column, _ in required_fields])
    rows_used = active_rows(ws, rows, trigger_columns)

    rows_to_validate = rows_used
    if require_one_row and not rows_used:
        rows_to_validate = [rows[0]]

    for sequential_number, row in enumerate(rows_to_validate, start=1):
        row_label = f"{row_name} {sequential_number} (fila {row})"
        for column, field in required_fields:
            cell = f"{column}{row}"
            if not is_filled(get_value(ws, cell)):
                add_missing(missing, section, field, cell, row_label=row_label)

    return rows_used


def validate_signature(
    ws,
    missing: List[Dict[str, str]],
    field: str,
    name_cell: str,
    id_cell: str,
) -> None:
    missing_parts = []
    if not is_filled(get_value(ws, name_cell)):
        missing_parts.append("nombre")
    if not is_filled(get_value(ws, id_cell)):
        missing_parts.append("RUT")

    if missing_parts:
        add_missing(
            missing,
            "Firmas y validación",
            field,
            f"{name_cell}/{id_cell}",
            f"Falta {' y '.join(missing_parts)} en el recuadro de firma",
            priority=PRIORITY_CRITICAL,
        )


def get_ot_worksheet(workbook):
    preferred_names = ["OT FORMATO IMPRIMIR", "OT", "ORDEN DE TRABAJO"]
    normalized = {name.strip().upper(): name for name in workbook.sheetnames}
    for preferred in preferred_names:
        if preferred in normalized:
            return workbook[normalized[preferred]]
    return workbook[workbook.sheetnames[0]]


def validate_work_order(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    workbook = load_workbook(
        io.BytesIO(file_bytes),
        data_only=True,
        read_only=False,
        keep_vba=filename.lower().endswith(".xlsm"),
    )
    ws = get_ot_worksheet(workbook)

    missing: List[Dict[str, str]] = []

    # ---------------------------------------------------------
    # CAMPOS ESENCIALES DEFINIDOS PARA LA VALIDACIÓN DE LA OT
    # ---------------------------------------------------------
    # 1) Horómetro. Debe contener información y ser distinto de cero.
    horometer_cell = "G13"
    horometer_value = get_value(ws, horometer_cell)

    if not is_filled(horometer_value):
        add_missing(
            missing,
            "Información general",
            "Horómetro",
            horometer_cell,
            "Campo requerido sin información",
        )
    elif is_zero_numeric(horometer_value):
        add_missing(
            missing,
            "Información general",
            "Horómetro",
            horometer_cell,
            "Valor 0: se considera falta de información en el horómetro",
        )

    # 2) Motivo de detención del equipo.
    validate_single_cells(
        ws,
        missing,
        "Motivo de detención",
        [("Motivo detención del equipo", "AB25")],
    )

    # 3) Por cada fila de actividad utilizada se exige únicamente:
    #    descripción del síntoma y códigos de trabajo, síntoma y causa.
    # Las demás columnas solo ayudan a reconocer que una fila está en uso.
    work_rows_used = validate_required_table(
        ws,
        missing,
        "Información del trabajo",
        WORK_ROWS,
        ESSENTIAL_WORK_FIELDS,
        trigger_columns=[column for column, _ in WORK_FIELDS],
        require_one_row=True,
        row_name="Actividad",
    )

    # El código causa 6.6 o 7.1 corresponde a la categoría "Otros" y no es válido.
    rows_to_check = work_rows_used or [WORK_ROWS[0]]
    for sequential_number, row in enumerate(rows_to_check, start=1):
        cell = f"BH{row}"
        value = get_value(ws, cell)
        if is_invalid_cause_code(value):
            add_missing(
                missing,
                "Información del trabajo",
                "Código causa",
                cell,
                f"Valor inválido {normalize_value(value)}: los códigos 6.6 y 7.1 corresponden a la categoría 'Otros'",
                row_label=f"Actividad {sequential_number} (fila {row})",
                priority=PRIORITY_CRITICAL,
            )

    # 4) Descripción de actividades: basta con que exista información en una de sus líneas.
    activity_description_cells = [f"B{row}" for row in range(98, 125, 3)]
    if not any(is_filled(get_value(ws, cell)) for cell in activity_description_cells):
        add_missing(
            missing,
            "Descripción de actividades",
            "Descripción de actividades",
            "B98:BZ124",
        )

    # 5) Firmas del reverso: cada recuadro debe contener nombre y RUT.
    validate_signature(
        ws,
        missing,
        "Firma jefe turno (nombre + RUT)",
        "C238",
        "C244",
    )
    validate_signature(
        ws,
        missing,
        "Firma técnico responsable (nombre + RUT)",
        "BD239",
        "BD243",
    )

    equipo = first_filled(ws, ["G7"])
    orden = first_filled(ws, ["G25", "J42", "Z145"])
    turno = first_filled(ws, ["G19"])
    critical_missing = sum(1 for item in missing if item["Prioridad"] == PRIORITY_CRITICAL)

    return {
        "Archivo": filename,
        "Equipo": equipo,
        "Orden": orden,
        "Turno": turno,
        "Campos faltantes": len(missing),
        "Faltantes críticos": critical_missing,
        "Estado": "Completa" if not missing else "Con observaciones",
        "Campos con observación": summarize_observations(missing),
        "missing": missing,
    }


# ============================================================
# Lectura de archivos subidos
# ============================================================
def read_uploaded_files(uploaded_files) -> List[Tuple[str, bytes]]:
    files: List[Tuple[str, bytes]] = []
    for uploaded in uploaded_files:
        name = uploaded.name
        data = uploaded.getvalue()
        lower = name.lower()

        if lower.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(data)) as zip_file:
                for info in zip_file.infolist():
                    internal_name = info.filename.replace("\\", "/")
                    if (
                        internal_name.lower().endswith((".xlsx", ".xlsm"))
                        and not internal_name.startswith("__MACOSX")
                        and not internal_name.split("/")[-1].startswith("~$")
                    ):
                        files.append((internal_name.split("/")[-1], zip_file.read(info)))
        elif lower.endswith((".xlsx", ".xlsm")) and not name.startswith("~$"):
            files.append((name, data))

    return files


# ============================================================
# Construcción de tablas de resumen
# ============================================================
def create_result_dataframes(results: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    summary_records: List[Dict[str, Any]] = []
    detail_records: List[Dict[str, Any]] = []

    for result in results:
        summary_records.append({key: result[key] for key in SUMMARY_COLUMNS})
        for item in result["missing"]:
            detail_records.append(
                {
                    "Archivo": result["Archivo"],
                    "Equipo": result["Equipo"],
                    "Orden": result["Orden"],
                    **item,
                }
            )

    summary_df = pd.DataFrame(summary_records, columns=SUMMARY_COLUMNS)
    detail_df = pd.DataFrame(detail_records, columns=DETAIL_COLUMNS)
    return summary_df, detail_df


def build_field_summary(detail_df: pd.DataFrame, total_orders: int) -> pd.DataFrame:
    columns = ["Prioridad", "Campo faltante", "Cantidad faltante", "OT afectadas", "% OT afectadas"]
    if detail_df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        detail_df.groupby(["Prioridad", "Campo faltante"], as_index=False)
        .agg(
            **{
                "Cantidad faltante": ("Campo faltante", "size"),
                "OT afectadas": ("Archivo", "nunique"),
            }
        )
    )
    summary["% OT afectadas"] = (
        summary["OT afectadas"] / max(total_orders, 1) * 100
    ).round(1)
    priority_order = {PRIORITY_CRITICAL: 0, PRIORITY_STANDARD: 1}
    summary["_priority_order"] = summary["Prioridad"].map(priority_order).fillna(9)
    summary = summary.sort_values(
        ["_priority_order", "Cantidad faltante", "Campo faltante"],
        ascending=[True, False, True],
    ).drop(columns="_priority_order")
    return summary[columns]


def build_compliance_summary(detail_df: pd.DataFrame, total_orders: int) -> pd.DataFrame:
    """Resume el porcentaje de OT conformes y con observación para cada campo validado."""
    records: List[Dict[str, Any]] = []

    if total_orders <= 0:
        return pd.DataFrame(
            columns=["Campo", "OT conformes", "OT con observación", "% Cumplimiento", "% Con observación"]
        )

    affected_by_field: Dict[str, int] = {}
    if not detail_df.empty:
        affected_by_field = (
            detail_df.groupby("Campo faltante")["Archivo"]
            .nunique()
            .astype(int)
            .to_dict()
        )

    for field in VALIDATED_FIELDS:
        affected = min(affected_by_field.get(field, 0), total_orders)
        conforming = total_orders - affected
        records.append(
            {
                "Campo": field,
                "OT conformes": conforming,
                "OT con observación": affected,
                "% Cumplimiento": round(conforming / total_orders * 100, 1),
                "% Con observación": round(affected / total_orders * 100, 1),
            }
        )

    return pd.DataFrame(records)


def build_section_summary(detail_df: pd.DataFrame) -> pd.DataFrame:
    columns = ["Sección", "Campos faltantes", "% del total"]
    if detail_df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        detail_df.groupby("Sección", as_index=False)
        .size()
        .rename(columns={"size": "Campos faltantes"})
    )
    total_missing = int(summary["Campos faltantes"].sum())
    summary["% del total"] = (summary["Campos faltantes"] / total_missing * 100).round(1)
    return summary.sort_values("Campos faltantes", ascending=False)[columns]


# ============================================================
# Reporte Excel
# ============================================================
def build_excel_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    field_summary_df: pd.DataFrame,
    section_summary_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    total_orders = len(summary_df)
    orders_with_obs = int((summary_df["Estado"] == "Con observaciones").sum()) if total_orders else 0
    complete_orders = int((summary_df["Estado"] == "Completa").sum()) if total_orders else 0
    total_missing = int(summary_df["Campos faltantes"].sum()) if total_orders else 0
    critical_missing = int(summary_df["Faltantes críticos"].sum()) if total_orders else 0
    compliance_summary_df = build_compliance_summary(detail_df, total_orders)

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Resumen OT")
        compliance_summary_df.to_excel(writer, index=False, sheet_name="Cumplimiento por campo")
        field_summary_df.to_excel(writer, index=False, sheet_name="Resumen por campo")
        section_summary_df.to_excel(writer, index=False, sheet_name="Resumen secciones")
        detail_df.to_excel(writer, index=False, sheet_name="Detalle faltantes")

        workbook = writer.book
        header_fmt = workbook.add_format(
            {"bold": True, "bg_color": "#FFC400", "border": 1, "align": "center", "valign": "vcenter"}
        )
        body_fmt = workbook.add_format({"border": 1, "valign": "top"})
        bad_fmt = workbook.add_format({"bg_color": "#FFECEC", "font_color": "#C00000", "border": 1})
        critical_fmt = workbook.add_format({"bg_color": "#FECACA", "font_color": "#991B1B", "border": 1, "bold": True})
        ok_fmt = workbook.add_format({"bg_color": "#FFC400", "font_color": "#111111", "border": 1})
        title_fmt = workbook.add_format({"bold": True, "font_size": 18, "bg_color": "#111111", "font_color": "#FFFFFF", "align": "center"})
        kpi_title_fmt = workbook.add_format({"bold": True, "bg_color": "#FFC400", "border": 1, "align": "center"})
        kpi_value_fmt = workbook.add_format({"bold": True, "font_size": 20, "border": 1, "align": "center"})

        for sheet_name, dataframe in [
            ("Resumen OT", summary_df),
            ("Cumplimiento por campo", compliance_summary_df),
            ("Resumen por campo", field_summary_df),
            ("Resumen secciones", section_summary_df),
            ("Detalle faltantes", detail_df),
        ]:
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes(1, 0)
            if len(dataframe.columns) > 0:
                worksheet.autofilter(0, 0, max(len(dataframe), 1), len(dataframe.columns) - 1)
            for col_num, col_name in enumerate(dataframe.columns):
                worksheet.write(0, col_num, col_name, header_fmt)
                width = min(max(len(str(col_name)) + 4, 14), 45)
                if not dataframe.empty:
                    width = min(max(width, int(dataframe[col_name].astype(str).map(len).max()) + 3), 55)
                worksheet.set_column(col_num, col_num, width, body_fmt)

        if not summary_df.empty:
            ws_summary = writer.sheets["Resumen OT"]
            estado_col = summary_df.columns.get_loc("Estado")
            critical_col = summary_df.columns.get_loc("Faltantes críticos")
            ws_summary.conditional_format(
                1,
                estado_col,
                len(summary_df),
                estado_col,
                {"type": "text", "criteria": "containing", "value": "Con observaciones", "format": bad_fmt},
            )
            ws_summary.conditional_format(
                1,
                estado_col,
                len(summary_df),
                estado_col,
                {"type": "text", "criteria": "containing", "value": "Completa", "format": ok_fmt},
            )
            ws_summary.conditional_format(
                1,
                critical_col,
                len(summary_df),
                critical_col,
                {"type": "cell", "criteria": ">", "value": 0, "format": critical_fmt},
            )

        if not detail_df.empty:
            ws_detail = writer.sheets["Detalle faltantes"]
            priority_col = detail_df.columns.get_loc("Prioridad")
            ws_detail.conditional_format(
                1,
                priority_col,
                len(detail_df),
                priority_col,
                {"type": "text", "criteria": "containing", "value": PRIORITY_CRITICAL, "format": critical_fmt},
            )

        dashboard = workbook.add_worksheet("Dashboard")
        writer.sheets["Dashboard"] = dashboard
        dashboard.hide_gridlines(2)
        dashboard.set_column("A:A", 3)
        dashboard.set_column("B:I", 18)
        dashboard.merge_range("B2:I3", "REPORTE EJECUTIVO DE VALIDACIÓN OT", title_fmt)

        kpis = [
            ("Órdenes procesadas", total_orders),
            ("Con observaciones", orders_with_obs),
            ("Campos faltantes", total_missing),
            ("Faltantes críticos", critical_missing),
            ("Órdenes completas", complete_orders),
        ]
        for index, (title, value) in enumerate(kpis):
            col = 1 + index
            dashboard.write(4, col, title, kpi_title_fmt)
            dashboard.write(5, col, value, kpi_value_fmt)

        if not compliance_summary_df.empty:
            chart = workbook.add_chart({"type": "bar", "subtype": "stacked"})
            last_row = len(compliance_summary_df)
            chart.add_series(
                {
                    "name": "Conforme",
                    "categories": ["Cumplimiento por campo", 1, 0, last_row, 0],
                    "values": ["Cumplimiento por campo", 1, 3, last_row, 3],
                    "fill": {"color": "#FFC400"},
                    "border": {"color": "#D6A400"},
                    "data_labels": {"value": True, "num_format": r"0.0\%"},
                }
            )
            chart.add_series(
                {
                    "name": "Con observación",
                    "categories": ["Cumplimiento por campo", 1, 0, last_row, 0],
                    "values": ["Cumplimiento por campo", 1, 4, last_row, 4],
                    "fill": {"color": "#EF4444"},
                    "border": {"color": "#B91C1C"},
                    "data_labels": {"value": True, "num_format": r"0.0\%"},
                }
            )
            chart.set_title({"name": "Cumplimiento por campo"})
            chart.set_legend({"position": "bottom"})
            chart.set_x_axis({"name": "Porcentaje de OT", "min": 0, "max": 100, "major_unit": 20})
            chart.set_y_axis({"reverse": True})
            chart.set_style(10)
            dashboard.insert_chart("B9", chart, {"x_scale": 1.5, "y_scale": 1.45})

        status_data_row = 32
        dashboard.write(status_data_row, 1, "Estado", header_fmt)
        dashboard.write(status_data_row, 2, "Cantidad", header_fmt)
        dashboard.write(status_data_row + 1, 1, "Completas", body_fmt)
        dashboard.write(status_data_row + 1, 2, complete_orders, body_fmt)
        dashboard.write(status_data_row + 2, 1, "Con observaciones", body_fmt)
        dashboard.write(status_data_row + 2, 2, orders_with_obs, body_fmt)

        if total_orders:
            donut = workbook.add_chart({"type": "doughnut"})
            donut.add_series(
                {
                    "name": "Estado de órdenes",
                    "categories": ["Dashboard", status_data_row + 1, 1, status_data_row + 2, 1],
                    "values": ["Dashboard", status_data_row + 1, 2, status_data_row + 2, 2],
                    "points": [
                        {"fill": {"color": "#FFC400"}},
                        {"fill": {"color": "#EF4444"}},
                    ],
                    "data_labels": {"percentage": True, "category": True},
                }
            )
            donut.set_title({"name": "Estado de órdenes"})
            donut.set_hole_size(55)
            donut.set_legend({"position": "bottom"})
            dashboard.insert_chart("K9", donut, {"x_scale": 1.1, "y_scale": 1.2})

    return output.getvalue()


# ============================================================
# Reporte PDF
# ============================================================
def build_pdf_report(
    summary_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    field_summary_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Reporte Ejecutivo de Validación de Órdenes de Trabajo", styles["Title"]),
        Paragraph(f"Fecha reporte: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 0.3 * cm),
    ]

    total_orders = len(summary_df)
    orders_with_obs = int((summary_df["Estado"] == "Con observaciones").sum()) if total_orders else 0
    complete_orders = int((summary_df["Estado"] == "Completa").sum()) if total_orders else 0
    total_missing = int(summary_df["Campos faltantes"].sum()) if total_orders else 0
    critical_missing = int(summary_df["Faltantes críticos"].sum()) if total_orders else 0

    kpi_data = [
        ["Órdenes totales", "Con observaciones", "Campos faltantes", "Faltantes críticos", "Completas"],
        [total_orders, orders_with_obs, total_missing, critical_missing, complete_orders],
    ]
    kpi_table = Table(kpi_data, colWidths=[4.8 * cm] * 5)
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFC400")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, 1), 18),
                ("BACKGROUND", (4, 1), (4, 1), colors.HexColor("#FFC400")),
                ("TEXTCOLOR", (4, 1), (4, 1), colors.black),
            ]
        )
    )
    story.extend([kpi_table, Spacer(1, 0.4 * cm)])

    story.append(Paragraph("Resumen por OT", styles["Heading2"]))
    pdf_summary_columns = ["Equipo", "Orden", "Estado", "Campos faltantes"]
    pdf_summary = summary_df[pdf_summary_columns].head(20).copy()
    pdf_summary_data = [pdf_summary.columns.tolist()] + pdf_summary.astype(str).values.tolist()
    pdf_summary_table = Table(pdf_summary_data, repeatRows=1, colWidths=[4.0 * cm, 4.0 * cm, 5.0 * cm, 4.0 * cm])
    pdf_summary_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
    ]
    for row_number, estado in enumerate(pdf_summary["Estado"].tolist(), start=1):
        if estado == "Completa":
            pdf_summary_style.append(("BACKGROUND", (0, row_number), (-1, row_number), colors.HexColor("#FFC400")))
            pdf_summary_style.append(("TEXTCOLOR", (0, row_number), (-1, row_number), colors.black))
        else:
            pdf_summary_style.append(("BACKGROUND", (0, row_number), (-1, row_number), colors.HexColor("#FFECEC")))
            pdf_summary_style.append(("TEXTCOLOR", (0, row_number), (-1, row_number), colors.HexColor("#991B1B")))
    pdf_summary_table.setStyle(TableStyle(pdf_summary_style))
    story.extend([pdf_summary_table, Spacer(1, 0.4 * cm)])

    story.append(Paragraph("Campos faltantes por campo", styles["Heading2"]))
    field_show = field_summary_df.head(15).copy()
    field_table_data = [field_show.columns.tolist()] + field_show.astype(str).values.tolist()
    field_table = Table(field_table_data, repeatRows=1)
    field_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFC400")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#FFF7D6")),
            ]
        )
    )
    story.extend([field_table, Spacer(1, 0.4 * cm)])

    story.append(Paragraph("Detalle de órdenes con observaciones", styles["Heading2"]))
    detail_columns = [
        "Archivo",
        "Equipo",
        "Orden",
        "Sección",
        "Campo faltante",
        "Prioridad",
        "Celda/Rango",
    ]
    detail_show = (
        detail_df[detail_columns].head(30).copy()
        if not detail_df.empty
        else pd.DataFrame(columns=detail_columns)
    )
    detail_table_data = [detail_show.columns.tolist()] + detail_show.astype(str).values.tolist()
    detail_table = Table(detail_table_data, repeatRows=1)
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFC400")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(detail_table)

    document.build(story)
    return output.getvalue()


# ============================================================
# Sidebar
# ============================================================
with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
    else:
        st.warning("No se encontró el archivo Finning-CAT.png en la carpeta del proyecto.")
    st.markdown("## VALIDACIÓN\n## ÓRDENES DE TRABAJO")
    st.markdown("---")
    st.markdown(
        '<div class="step-card"><span class="step-number">1</span><b>Subir archivos Excel</b><br><small>Formatos: .xlsx, .xlsm o .zip con varias OT.</small></div>',
        unsafe_allow_html=True,
    )
    uploaded_files = st.file_uploader(
        "Arrastra y suelta tus órdenes aquí",
        type=["xlsx", "xlsm", "zip"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    st.markdown(
        '<div class="step-card"><span class="step-number">2</span><b>Validar órdenes</b><br><small>Se revisan únicamente los campos esenciales definidos.</small></div>',
        unsafe_allow_html=True,
    )
    validate_btn = st.button("▶️ Validar órdenes", use_container_width=True)
    st.markdown(
        '<div class="step-card"><span class="step-number">3</span><b>Descargar reporte</b><br><small>Se generan reportes en Excel y PDF.</small></div>',
        unsafe_allow_html=True,
    )
    st.caption("Prioridad crítica: código trabajo, síntoma, causa y firmas. Causa 6.6 o 7.1 = inválida.")

# ============================================================
# Pantalla principal
# ============================================================
st.markdown('<p class="main-title">REPORTE EJECUTIVO DE VALIDACIÓN</p>', unsafe_allow_html=True)
st.markdown(
    '<p class="subtitle">Validación de horómetro, motivo de detención, síntoma, códigos principales, descripción de actividades y firmas de la Orden de Trabajo.</p>',
    unsafe_allow_html=True,
)
st.markdown(
    '<div class="critical-note"><b>Campos revisados:</b> Horómetro, motivo de detención, descripción del síntoma, Código trabajo, Código síntoma, Código causa, descripción de actividades y firmas. <b>Críticos:</b> los tres códigos y ambas firmas. Los códigos causa <b>6.6</b> y <b>7.1</b> se consideran inválidos.</div>',
    unsafe_allow_html=True,
)

if "results" not in st.session_state:
    st.session_state.results = None

if validate_btn:
    if not uploaded_files:
        st.warning("Debes subir al menos un archivo Excel o ZIP para validar.")
    else:
        files = read_uploaded_files(uploaded_files)
        if not files:
            st.error("No se encontraron archivos .xlsx o .xlsm válidos.")
        else:
            progress = st.progress(0, text="Validando órdenes...")
            results: List[Dict[str, Any]] = []
            errors: List[Dict[str, str]] = []

            for index, (filename, data) in enumerate(files, start=1):
                try:
                    results.append(validate_work_order(data, filename))
                except Exception as error:
                    errors.append({"Archivo": filename, "Error": str(error)})
                progress.progress(
                    index / len(files),
                    text=f"Validando {index} de {len(files)}: {filename}",
                )

            progress.empty()
            st.session_state.results = {"results": results, "errors": errors}

if st.session_state.results is None:
    st.info("Sube una o varias órdenes de trabajo y presiona *Validar órdenes* para generar el reporte.")
    st.stop()

results = st.session_state.results["results"]
errors = st.session_state.results["errors"]

if not results:
    st.error("No fue posible procesar ninguna orden de trabajo.")
    if errors:
        st.dataframe(pd.DataFrame(errors), use_container_width=True, hide_index=True)
    st.stop()

summary_df, detail_df = create_result_dataframes(results)
total_orders = len(summary_df)
field_summary_df = build_field_summary(detail_df, total_orders)
compliance_summary_df = build_compliance_summary(detail_df, total_orders)
section_summary_df = build_section_summary(detail_df)

orders_with_obs = int((summary_df["Estado"] == "Con observaciones").sum())
complete_orders = int((summary_df["Estado"] == "Completa").sum())
total_missing = int(summary_df["Campos faltantes"].sum())
critical_missing = int(summary_df["Faltantes críticos"].sum())

# KPIs
kpi_columns = st.columns(5)
kpis = [
    ("Órdenes totales", total_orders, "Órdenes procesadas"),
    ("Con observaciones", orders_with_obs, "OT con al menos un hallazgo"),
    ("Campos faltantes", total_missing, "Incluye valores inválidos"),
    ("Faltantes críticos", critical_missing, "Códigos o firmas"),
    ("Órdenes completas", complete_orders, "Sin campos faltantes"),
]
for column, (title, value, note) in zip(kpi_columns, kpis):
    with column:
        st.markdown(
            f'<div class="kpi-card"><div class="kpi-title">{title}</div><div class="kpi-value">{value}</div><div class="kpi-note">{note}</div></div>',
            unsafe_allow_html=True,
        )

st.write("")

# Gráficos
left_chart, right_chart = st.columns([1.45, 1])
with left_chart:
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    st.subheader("Cumplimiento por campo")
    st.caption("Cada barra representa el 100% de las OT procesadas y facilita comparar todos los campos, incluso cuando existe un solo hallazgo.")

    chart_df = compliance_summary_df.melt(
        id_vars=["Campo"],
        value_vars=["% Cumplimiento", "% Con observación"],
        var_name="Resultado",
        value_name="Porcentaje",
    )
    chart_df["Resultado"] = chart_df["Resultado"].replace(
        {"% Cumplimiento": "Conforme", "% Con observación": "Con observación"}
    )
    chart_df["Etiqueta"] = chart_df["Porcentaje"].apply(
        lambda value: f"{value:.1f}%" if value >= 4 else ""
    )

    figure = px.bar(
        chart_df,
        x="Porcentaje",
        y="Campo",
        orientation="h",
        color="Resultado",
        text="Etiqueta",
        barmode="stack",
        color_discrete_map={"Conforme": OK_COLOR, "Con observación": OBS_COLOR},
        category_orders={
            "Campo": list(reversed(VALIDATED_FIELDS)),
            "Resultado": ["Conforme", "Con observación"],
        },
        template="plotly_white",
    )
    figure.update_traces(textposition="inside", insidetextanchor="middle")
    figure.update_layout(
        height=max(430, 42 * len(VALIDATED_FIELDS)),
        margin=dict(l=10, r=20, t=10, b=10),
        xaxis_title="Porcentaje de OT",
        yaxis_title="",
        xaxis=dict(range=[0, 100], ticksuffix="%"),
        legend_title_text="Resultado",
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(color="#111111"),
    )
    st.plotly_chart(figure, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)

with right_chart:
    st.markdown('<div class="panel">', unsafe_allow_html=True)
    st.subheader("Estado de órdenes")
    status_df = pd.DataFrame(
        {
            "Estado": ["Completas", "Con observaciones"],
            "Cantidad": [complete_orders, orders_with_obs],
        }
    )
    figure = px.pie(
        status_df,
        values="Cantidad",
        names="Estado",
        hole=0.55,
        color="Estado",
        color_discrete_map={"Completas": OK_COLOR, "Con observaciones": OBS_COLOR},
        template="plotly_white",
    )
    figure.update_layout(
        margin=dict(l=10, r=10, t=10, b=10),
        height=360,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(color="#111111"),
    )
    st.plotly_chart(figure, use_container_width=True)
    completion_percentage = complete_orders / total_orders * 100 if total_orders else 0
    st.info(
        f"{complete_orders} de {total_orders} órdenes ({completion_percentage:.1f}%) están completas."
    )
    st.markdown('</div>', unsafe_allow_html=True)

st.write("")

# Resúmenes
summary_tab, field_tab, detail_tab = st.tabs(
    ["Resumen por OT", "Resumen de campos faltantes", "Detalle de hallazgos"]
)

with summary_tab:
    st.dataframe(
        summary_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Campos con observación": st.column_config.TextColumn(
                "Campos con observación",
                help="Identifica los campos faltantes o inválidos detectados en cada OT.",
                width="large",
            )
        },
    )

with field_tab:
    if field_summary_df.empty:
        st.success("No existen campos faltantes para resumir.")
    else:
        st.dataframe(field_summary_df, use_container_width=True, hide_index=True)
        st.markdown("#### Campos faltantes por sección")
        st.dataframe(section_summary_df, use_container_width=True, hide_index=True)

with detail_tab:
    if detail_df.empty:
        st.success("Todas las órdenes revisadas están completas.")
    else:
        selected_priority = st.multiselect(
            "Filtrar por prioridad",
            options=[PRIORITY_CRITICAL, PRIORITY_STANDARD],
            default=[PRIORITY_CRITICAL, PRIORITY_STANDARD],
        )
        filtered_detail = detail_df[detail_df["Prioridad"].isin(selected_priority)]
        st.dataframe(filtered_detail, use_container_width=True, hide_index=True)

if errors:
    with st.expander("Archivos no procesados"):
        st.dataframe(pd.DataFrame(errors), use_container_width=True, hide_index=True)

excel_bytes = build_excel_report(
    summary_df,
    detail_df,
    field_summary_df,
    section_summary_df,
)
pdf_bytes = build_pdf_report(summary_df, detail_df, field_summary_df)

download_excel, download_pdf, _ = st.columns([1, 1, 2])
with download_excel:
    st.download_button(
        "⬇️ Descargar Excel",
        data=excel_bytes,
        file_name=f"reporte_validacion_ot_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with download_pdf:
    st.download_button(
        "⬇️ Descargar PDF",
        data=pdf_bytes,
        file_name=f"reporte_ejecutivo_ot_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

with st.expander("Campos revisados en el formato OT"):
    st.markdown(
        """
        La validación considera exclusivamente los siguientes campos:

        - *Horómetro; debe contener un valor distinto de **0*.
        - *Motivo de detención del equipo*.
        - *Descripción del síntoma*.
        - *Código trabajo*.
        - *Código síntoma*.
        - *Código causa; los valores **6.6* y *7.1* se consideran inválidos porque corresponden a la categoría “Otros”.
        - *Descripción de actividades*.
        - *Firma del jefe de turno*, verificando nombre y RUT.
        - *Firma del técnico responsable*, verificando nombre y RUT.

        No se generan observaciones por ningún otro campo del anverso o reverso.
        """
    )
